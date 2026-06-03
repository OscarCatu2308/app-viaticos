# ============================================================
#  APP VIÁTICOS — Descarga diaria + sube a Google Sheets
#  Aprende (NIT 30613655)
#  Descarga XMLs del día anterior
#  Parsea y sube a hoja "Documentos" en Google Sheets
#  Genera CSV y Excel de control
# ============================================================

from playwright.sync_api import sync_playwright
from datetime import date, timedelta, datetime
import calendar
import os
import zipfile
import xml.etree.ElementTree as ET
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter
import unicodedata
import gspread
from google.oauth2.service_account import Credentials

# ─────────────────────────────────────────────────────
# CONFIGURACION
# ─────────────────────────────────────────────────────
CLIENTE = {
    "nit"     : "30613655",
    "nombre"  : "Aprende",
    "password": "aprende",
}

CONFIG = {
    "llave_proveedor" : "Cchica",
    "descripcion"     : "Gastos Varios",
    "carpeta_base"    : os.path.join(
                            os.path.expanduser("~"),
                            "Desktop", "agente-sat", "app_viaticos"
                        ),
    "correo_remitente": "oscar.catu@asprobpo.com",
    "correo_destinos" : ["oscar.catu@asprobpo.com"],
    "credentials_file": os.path.join(
                            os.path.expanduser("~"),
                            "Desktop", "agente-sat", "google_credentials.json"
                        ),
    "sheet_id"        : "13gJPXnQMaZNj5qFiUBUrZR8yiqG-w_8hWDmp-dW0jZI",
}

NS_LIST = [
    {"dte": "http://www.sat.gob.gt/dte/fel/0.2.0"},
    {"dte": "http://www.sat.gob.gt/dte/fel/0.1.0"},
]

PALABRAS_COMBUSTIBLE = [
    "idp","impuesto al petroleo","impuesto petroleo",
    "distribucion de petroleo","regular","super","diesel",
    "premium","gasolina","combustible","galones"
]

CODIGOS_IDP = ["idp","imp. petroleo","impuesto petroleo",
               "impuesto al petroleo","petroleo","i.d.p"]

TIPOS_PEQUE = ["FPEQ","FESP"]

# ─────────────────────────────────────────────────────
# FECHAS — DIA ANTERIOR
# ─────────────────────────────────────────────────────
def obtener_dia_anterior():
    ayer   = date.today() - timedelta(days=1)
    return ayer, ayer, ayer.strftime("%Y-%m-%d")

# ─────────────────────────────────────────────────────
# LIMPIEZA
# ─────────────────────────────────────────────────────
def limpiar(valor):
    if not valor: return ""
    texto = unicodedata.normalize("NFD", str(valor))
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    for car in [",",";",'"',"'","|","\n","\r","\t"]:
        texto = texto.replace(car, " ")
    while "  " in texto: texto = texto.replace("  ", " ")
    return texto.strip()

def limpiar_descripcion(desc):
    if not desc: return ""
    desc = desc.strip()
    if "|" in desc:
        partes = [p.strip() for p in desc.split("|")]
        primera = partes[0].strip()
        if primera.isdigit() or (len(primera)<20 and
                primera.replace("-","").replace("_","").isalnum()):
            desc = "|".join(partes[1:]).strip()
    return limpiar(desc[:60])

# ─────────────────────────────────────────────────────
# GOOGLE SHEETS
# ─────────────────────────────────────────────────────
def conectar_sheets():
    try:
        return gspread.service_account(filename=CONFIG["credentials_file"])
    except Exception as e:
        print(f"  ⚠️  Error Google Sheets: {e}")
        return None

def subir_documentos_sheets(filas_docs):
    """Sube los documentos procesados a la hoja 'Documentos' en Google Sheets"""
    try:
        client = conectar_sheets()
        if not client:
            return False
        libro = client.open_by_key(CONFIG["sheet_id"])
        hojas = [h.title for h in libro.worksheets()]

        if "Documentos" in hojas:
            hoja = libro.worksheet("Documentos")
        else:
            hoja = libro.add_worksheet(title="Documentos", rows=5000, cols=25)

        # Limpiar y reescribir
        hoja.clear()
        encabezados = [
            "Serie","Correlativo","Referencia","Fecha","Hora",
            "NIT","Emisor","Departamento","Municipio","Direccion",
            "MontoTotal","MontoIVA","MontoExento","IVADeducible",
            "TipoTransaccion","DescCorta","EsCombustible",
            "TipoComb","Galones","PrecioGalon","Periodo"
        ]
        hoja.append_row(encabezados)

        filas_upload = []
        for d in filas_docs:
            filas_upload.append([
                d["serie"], d["correlativo"], d["referencia"],
                d["fecha"], d["hora"], d["nit"], d["emisor"],
                d["departamento"], d["municipio"], d["direccion"],
                d["gran_total"], d["monto_iva"], d["monto_exento"],
                d["iva_ded"], d["tipo_trans"], d["desc_corta"],
                "Si" if d["es_combustible"] else "No",
                d["tipo_comb"], d["galones"], d["precio_galon"],
                d.get("periodo","")
            ])

        if filas_upload:
            hoja.append_rows(filas_upload)

        print(f"  ✅ {len(filas_upload)} documentos subidos a Google Sheets (hoja Documentos)")
        return True
    except Exception as e:
        print(f"  ⚠️  Error subiendo a Sheets: {e}")
        return False

# ─────────────────────────────────────────────────────
# SAT — PLAYWRIGHT
# ─────────────────────────────────────────────────────
def login(pagina, nit, password):
    pagina.goto("https://farm3.sat.gob.gt/menu/login.jsf")
    pagina.wait_for_load_state("networkidle")
    pagina.wait_for_selector("input[placeholder='Usuario']", timeout=10000)
    pagina.fill("input[placeholder='Usuario']", nit)
    pagina.fill("input[placeholder='Contraseña']", password)
    pagina.click("button:has-text('INICIAR SESIÓN')")
    pagina.wait_for_url(lambda url: "login" not in url, timeout=30000)
    pagina.wait_for_load_state("networkidle")
    pagina.wait_for_timeout(3000)

def abrir_consulta_dte(pagina):
    pagina.click(".menu-toggle,[class*='menu'],[id*='menu'],button[class*='nav']")
    pagina.wait_for_timeout(1500)
    pagina.hover("text=Servicios Tributarios")
    pagina.wait_for_timeout(1000)
    pagina.hover("text=Factura Electrónica en Línea (FEL)")
    pagina.wait_for_timeout(1000)
    pagina.click("text=Consultar DTE")
    pagina.wait_for_load_state("networkidle")
    pagina.wait_for_timeout(5000)

def obtener_iframe(pagina):
    for frame in pagina.frames:
        if "felcons" in frame.url or "dte" in frame.url.lower():
            return frame
    return pagina.frames[1] if len(pagina.frames) > 1 else None

def cerrar_overlays(pagina, frame):
    try: pagina.keyboard.press("Escape"); pagina.wait_for_timeout(500)
    except: pass
    try:
        frame.evaluate("""() => { document.querySelectorAll(
            '.cdk-overlay-backdrop,.cdk-overlay-transparent-backdrop'
        ).forEach(b => b.click()); }""")
        pagina.wait_for_timeout(500)
    except: pass

def navegar_mes_calendario(frame, mes_obj, anio_obj):
    meses = {"ENE":1,"FEB":2,"MAR":3,"ABR":4,"MAY":5,"JUN":6,
             "JUL":7,"AGO":8,"SEP":9,"OCT":10,"NOV":11,"DIC":12}
    for _ in range(24):
        try: texto = frame.locator(".mat-calendar-period-button,.mat-calendar-header button").first.inner_text()
        except: break
        mes_actual  = next((v for k,v in meses.items() if k in texto.upper()), None)
        anio_actual = next((int(p) for p in texto.split() if p.isdigit() and len(p)==4), None)
        if mes_actual == mes_obj and anio_actual == anio_obj: break
        if mes_actual and anio_actual:
            if (anio_actual*12+mes_actual) > (anio_obj*12+mes_obj):
                frame.click("button.mat-calendar-previous-button,[aria-label='Previous month']")
            else:
                frame.click("button.mat-calendar-next-button,[aria-label='Next month']")
        else:
            frame.click("button.mat-calendar-previous-button,[aria-label='Previous month']")
        frame.wait_for_timeout(600)

def seleccionar_dia(frame, dia):
    frame.wait_for_timeout(500)
    celdas = frame.locator("td.mat-calendar-body-cell,.mat-calendar-body-cell")
    for i in range(celdas.count()):
        if celdas.nth(i).inner_text().strip() == str(dia):
            celdas.nth(i).click(); return

def abrir_calendario(frame, nombre_campo):
    frame.locator(f"input[name='{nombre_campo}']").click()
    frame.wait_for_timeout(500)
    if not frame.locator(".mat-calendar").is_visible():
        botones = frame.locator("mat-datepicker-toggle button,button.mat-icon-button")
        idx = 0 if nombre_campo == "fechaEmisionIni" else 1
        botones.nth(min(idx, botones.count()-1)).click()
    frame.wait_for_timeout(1000)

def llenar_fecha(frame, pagina, campo, fecha_obj):
    abrir_calendario(frame, campo)
    navegar_mes_calendario(frame, fecha_obj.month, fecha_obj.year)
    seleccionar_dia(frame, fecha_obj.day)
    frame.wait_for_timeout(500)
    pagina.keyboard.press("Escape")
    frame.wait_for_timeout(300)

def esta_marcado(frame):
    return frame.evaluate("""() => {
        const cb = document.querySelector('mat-header-cell mat-checkbox,th mat-checkbox');
        if (!cb) return false;
        return cb.classList.contains('mat-checkbox-checked') ||
               cb.getAttribute('ng-reflect-checked') === 'true';
    }""")

def seleccionar_todos(frame):
    frame.evaluate("() => window.scrollTo(0, 0)")
    frame.wait_for_timeout(500)
    for sel in ["mat-header-cell mat-checkbox .mat-checkbox-inner-container",
                "mat-header-cell mat-checkbox label","mat-header-cell mat-checkbox"]:
        try:
            el = frame.locator(sel).first
            if el.count() > 0:
                el.scroll_into_view_if_needed(); frame.wait_for_timeout(300)
                el.click(force=True); frame.wait_for_timeout(2000)
                if esta_marcado(frame): return True
        except: pass
    frame.evaluate("""() => {
        const cb = document.querySelector(
            'mat-header-cell mat-checkbox .mat-checkbox-inner-container');
        if (cb) cb.click();
    }""")
    frame.wait_for_timeout(2000)
    return esta_marcado(frame)

def extraer_zip(ruta_zip, carpeta_xml):
    count = 0
    try:
        with zipfile.ZipFile(ruta_zip, 'r') as z:
            for n in z.namelist():
                if n.lower().endswith(".xml"):
                    z.extract(n, carpeta_xml); count += 1
        os.remove(ruta_zip)
    except Exception as e:
        print(f"  ⚠️  Error ZIP: {e}")
    return count

def descargar_xmls(pagina, frame, fecha_ini, fecha_fin, carpeta):
    carpeta_xml = os.path.join(carpeta, "xml")
    os.makedirs(carpeta_xml, exist_ok=True)
    xmls = sum(len([f for f in files if f.lower().endswith(".xml")])
               for _, _, files in os.walk(carpeta_xml))
    if xmls > 0:
        print(f"  ℹ️  {xmls} XMLs ya descargados")
        return xmls
    cerrar_overlays(pagina, frame)
    frame.wait_for_timeout(1000)
    frame.click("button:has-text('Limpiar')")
    frame.wait_for_timeout(1500)
    frame.click("#mat-select-0"); frame.wait_for_timeout(1000)
    frame.click("mat-option:has-text('Recibidos')"); frame.wait_for_timeout(500)
    cerrar_overlays(pagina, frame)
    llenar_fecha(frame, pagina, "fechaEmisionIni",   fecha_ini)
    llenar_fecha(frame, pagina, "fechaEmisionFinal", fecha_fin)
    frame.click("button:has-text('Buscar')")
    frame.wait_for_load_state("networkidle"); frame.wait_for_timeout(5000)
    if frame.locator(".mat-row,mat-row").count() == 0:
        print("  ⚠️  Sin resultados"); return 0
    try:
        texto = frame.locator(".mat-paginator-range-label").inner_text().strip()
        total = int(texto.replace(",","").split()[-1])
        print(f"  Total DTEs: {total}")
    except: pass
    if not seleccionar_todos(frame):
        print("  ⚠️  No se pudo seleccionar todos"); return 0
    ruta_zip = os.path.join(carpeta, "temp.zip")
    try:
        with pagina.expect_download(timeout=600000) as dl_info:
            frame.locator(".iconDownload fa.xml,.iconDownload label.iconDownloadText.xml").first.click(force=True)
        dl_info.value.save_as(ruta_zip)
        count = extraer_zip(ruta_zip, carpeta_xml)
        print(f"  ✅ {count} XMLs extraidos"); return count
    except Exception as e:
        print(f"  ⚠️  Error descarga: {e}"); return 0

# ─────────────────────────────────────────────────────
# PARSEAR XML
# ─────────────────────────────────────────────────────
def find_ns(root, path):
    for ns in NS_LIST:
        el = root.find(path, ns)
        if el is not None: return el, ns
    return None, None

def get_nc_imp(imp, ns):
    nc = imp.find("dte:NombreCorto", ns)
    if nc is not None and nc.text: return nc.text.strip().lower()
    return (imp.get("NombreCorto","") or "").lower()

def extraer_combustible_item(item, ns):
    desc_el  = item.find("dte:Descripcion", ns)
    cant_el  = item.find("dte:Cantidad", ns)
    prec_el  = item.find("dte:PrecioUnitario", ns)
    desc     = (desc_el.text or "") if desc_el is not None else ""
    galones  = 0.0; precio = 0.0; tipo_c = ""
    if cant_el is not None and cant_el.text:
        try: galones = round(float(cant_el.text), 4)
        except: pass
    if prec_el is not None and prec_el.text:
        try: precio = round(float(prec_el.text), 4)
        except: pass
    if "|" in desc:
        partes = [p.strip() for p in desc.split("|")]
        tipo_c = partes[0]
        try:
            if len(partes) > 2: precio = max(precio, float(partes[2]))
        except: pass
    else:
        tipo_c = desc[:20]
    t = tipo_c.lower()
    if "regular" in t: tipo_c = "Regular"
    elif "super" in t: tipo_c = "Super"
    elif "diesel" in t: tipo_c = "Diesel"
    elif "premium" in t: tipo_c = "Premium"
    return tipo_c, round(galones, 4), round(precio, 4)

def parsear_xml(ruta_xml, periodo):
    try:
        tree = ET.parse(ruta_xml)
        root = tree.getroot()
        _, ns = find_ns(root, ".//dte:DatosGenerales")
        if ns is None: return None
        dg      = root.find(".//dte:DatosGenerales", ns)
        num_aut = root.find(".//dte:NumeroAutorizacion", ns)
        emisor  = root.find(".//dte:Emisor", ns)
        totales = root.find(".//dte:Totales", ns)
        if dg is None or num_aut is None: return None
        tipo_dte   = dg.get("Tipo","FACT")
        fecha_hora = dg.get("FechaHoraEmision","")
        fecha = hora = ""
        try:
            dt    = datetime.strptime(fecha_hora[:19],"%Y-%m-%dT%H:%M:%S")
            fecha = dt.strftime("%d/%m/%Y"); hora = dt.strftime("%H:%M:%S")
        except:
            fecha = fecha_hora[:10]; hora = fecha_hora[11:19] if len(fecha_hora)>10 else ""
        serie       = num_aut.get("Serie","")
        correlativo = num_aut.get("Numero","")
        nit_emisor = nom_emisor = dir_emisor = mun_emisor = dep_emisor = ""
        if emisor is not None:
            nit_emisor = emisor.get("NITEmisor","") or emisor.get("Nit","")
            nom_emisor = emisor.get("NombreEmisor","") or emisor.get("NombreComercial","")
            dir_el = emisor.find("dte:DireccionEmisor", ns)
            if dir_el is not None:
                d = dir_el.find("dte:Direccion",    ns)
                m = dir_el.find("dte:Municipio",    ns)
                e = dir_el.find("dte:Departamento", ns)
                dir_emisor = d.text if d is not None else ""
                mun_emisor = m.text if m is not None else ""
                dep_emisor = e.text if e is not None else ""
        gran_total = monto_iva = monto_exento = 0.0
        if totales is not None:
            gt = totales.find("dte:GranTotal", ns)
            if gt is not None and gt.text: gran_total = round(float(gt.text),2)
            for timp in totales.findall("dte:TotalImpuestos/dte:TotalImpuesto", ns):
                nc_el = timp.find("dte:NombreCorto", ns)
                nc    = nc_el.text if nc_el is not None else timp.get("NombreCorto","")
                if (nc or "").upper() == "IVA":
                    try:
                        tm = timp.find("dte:TotalMontoImpuesto", ns)
                        v  = tm.text if tm is not None else timp.get("TotalMontoImpuesto",0)
                        monto_iva = round(float(v),2)
                    except: pass
        items   = root.findall(".//dte:Items/dte:Item", ns)
        es_comb = False; tipo_comb_str = ""; galones = precio_galon = 0.0
        for item in items:
            desc_el = item.find("dte:Descripcion", ns)
            desc    = (desc_el.text or "") if desc_el is not None else ""
            imps    = item.findall("dte:Impuestos/dte:Impuesto", ns)
            for imp in imps:
                if any(c in get_nc_imp(imp,ns) for c in CODIGOS_IDP):
                    es_comb = True; break
            if es_comb:
                tipo_comb_str, galones, precio_galon = extraer_combustible_item(item,ns); break
            if any(p in desc.lower() for p in PALABRAS_COMBUSTIBLE):
                es_comb = True
                tipo_comb_str, galones, precio_galon = extraer_combustible_item(item,ns); break
        tipo_trans = "C" if es_comb else ("M" if any(i.get("BienOServicio","B")=="B" for i in items) else "S")
        iva_ded    = "P" if tipo_dte in TIPOS_PEQUE else ("N" if monto_iva==0 else "S")
        desc_corta = ""
        if items:
            de = items[0].find("dte:Descripcion", ns)
            if de is not None and de.text: desc_corta = limpiar_descripcion(de.text)
        if es_comb: desc_corta = f"Combustible {tipo_comb_str}".strip()
        llave = "00-Combustible" if es_comb else CONFIG["llave_proveedor"]
        return {
            "serie":limpiar(serie),"correlativo":limpiar(correlativo),
            "referencia":limpiar(f"{serie}-{correlativo}"),
            "fecha":fecha,"hora":hora,
            "nit":limpiar(nit_emisor),"emisor":limpiar(nom_emisor),
            "departamento":limpiar(dep_emisor),"municipio":limpiar(mun_emisor),
            "direccion":limpiar(dir_emisor),
            "gran_total":gran_total,"monto_iva":monto_iva,"monto_exento":monto_exento,
            "iva_ded":iva_ded,"tipo_trans":tipo_trans,"desc_corta":desc_corta,
            "es_combustible":es_comb,"tipo_comb":tipo_comb_str,
            "galones":galones,"precio_galon":precio_galon,
            "llave_proveedor":llave,"periodo":periodo,
            "fecha_libro":""
        }
    except Exception as e:
        print(f"  ⚠️  Error XML {os.path.basename(ruta_xml)}: {e}")
        return None

def procesar_xmls(carpeta_xml, periodo):
    filas = []
    archivos = []
    for root_dir, _, files in os.walk(carpeta_xml):
        for f in files:
            if f.lower().endswith(".xml"):
                archivos.append(os.path.join(root_dir, f))
    print(f"  XMLs a procesar: {len(archivos)}")
    for ruta in archivos:
        fila = parsear_xml(ruta, periodo)
        if fila: filas.append(fila)
    filas.sort(key=lambda f: (
        datetime.strptime(f["fecha"],"%d/%m/%Y") if f["fecha"] else datetime.min))
    return filas

# ─────────────────────────────────────────────────────
# GENERAR CSV Y EXCEL
# ─────────────────────────────────────────────────────
def generar_csv(filas, ruta_csv):
    with open(ruta_csv,"w",newline="",encoding="latin-1") as f:
        w = csv.writer(f)
        for fila in filas:
            w.writerow([
                fila["fecha"],fila["serie"],fila["correlativo"],
                fila["referencia"],fila["llave_proveedor"],
                CONFIG["descripcion"],fila["nit"],fila["emisor"],
                fila["gran_total"],fila["monto_iva"],fila["monto_exento"],
                fila["iva_ded"],fila["tipo_trans"],fila["fecha_libro"],
                "","",""
            ])
    print(f"  ✅ CSV: {os.path.basename(ruta_csv)} ({len(filas)} registros)")

def generar_excel(filas, ruta_excel, titulo):
    cols = ["Fecha","Serie","Correlativo","Referencia","LlaveProveedor",
            "NIT","Emisor","Departamento","Municipio","MontoTotal",
            "MontoIVA","IVADeducible","TipoTransaccion","DescCorta",
            "EsCombustible","TipoComb","Galones","PrecioGalon"]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Viaticos"
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    t = ws["A1"]; t.value = titulo
    t.font = Font(bold=True,size=13,color="FFFFFF")
    t.fill = PatternFill("solid",fgColor="1F4E79")
    t.alignment = Alignment(horizontal="center",vertical="center")
    hf = PatternFill("solid",fgColor="2E75B6")
    for col,nombre in enumerate(cols,1):
        c = ws.cell(row=2,column=col,value=nombre)
        c.font=Font(bold=True,color="FFFFFF",size=10)
        c.fill=hf; c.alignment=Alignment(horizontal="center")
    for ri,fila in enumerate(filas,3):
        fp = PatternFill("solid",fgColor="DEEAF1") if ri%2==0 else PatternFill("solid",fgColor="FFFFFF")
        vals = [fila["fecha"],fila["serie"],fila["correlativo"],
                fila["referencia"],fila["llave_proveedor"],
                fila["nit"],fila["emisor"],fila["departamento"],
                fila["municipio"],fila["gran_total"],fila["monto_iva"],
                fila["iva_ded"],fila["tipo_trans"],fila["desc_corta"],
                "Si" if fila["es_combustible"] else "No",
                fila["tipo_comb"],fila["galones"],fila["precio_galon"]]
        for ci,val in enumerate(vals,1):
            c=ws.cell(row=ri,column=ci,value=val); c.fill=fp
    ws.freeze_panes="A3"
    wb.save(ruta_excel)
    print(f"  ✅ Excel: {os.path.basename(ruta_excel)} ({len(filas)} registros)")

# ─────────────────────────────────────────────────────
# CORREO VIA OUTLOOK
# ─────────────────────────────────────────────────────
def enviar_correo(ruta_excel, ruta_csv, filas, periodo):
    try:
        import win32com.client
        total = round(sum(f["gran_total"] for f in filas),2)
        tipos = Counter(f["tipo_trans"] for f in filas)
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail    = outlook.CreateItem(0)
        for cuenta in outlook.Session.Accounts:
            if CONFIG["correo_remitente"].lower() in cuenta.SmtpAddress.lower():
                mail.SendUsingAccount = cuenta; break
        mail.To      = "; ".join(CONFIG["correo_destinos"])
        mail.Subject = f"App Viaticos — {CLIENTE['nombre']} — {periodo}"
        mail.HTMLBody = f"""
        <html><body style="font-family:Arial;font-size:13px">
        <p>Descarga diaria de viáticos <b>{CLIENTE['nombre']}</b> — <b>{periodo}</b></p>
        <table border="1" cellpadding="6" style="border-collapse:collapse">
          <tr style="background:#1F4E79;color:white"><th>Detalle</th><th>Valor</th></tr>
          <tr><td>Período</td><td>{periodo}</td></tr>
          <tr><td>Total facturas</td><td>{len(filas)}</td></tr>
          <tr><td>Combustible (C)</td><td>{tipos.get('C',0)}</td></tr>
          <tr><td>Servicios (S)</td><td>{tipos.get('S',0)}</td></tr>
          <tr><td>Materiales (M)</td><td>{tipos.get('M',0)}</td></tr>
          <tr><td><b>Total Q</b></td><td><b>Q {total:,.2f}</b></td></tr>
        </table>
        <p><em>Documentos disponibles en la app de viáticos.</em></p>
        </body></html>"""
        mail.Attachments.Add(ruta_excel)
        mail.Attachments.Add(ruta_csv)
        mail.Send()
        print(f"  ✅ Correo enviado")
    except Exception as e:
        print(f"  ⚠️  Error correo: {e}")

# ─────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────
def main():
    primer_dia, ultimo_dia, periodo = obtener_dia_anterior()

    print("\n" + "="*55)
    print("  APP VIÁTICOS — Descarga Diaria")
    print("="*55)
    print(f"  Cliente : {CLIENTE['nombre']} (NIT: {CLIENTE['nit']})")
    print(f"  Período : {periodo} (ayer)")
    print("="*55)

    carpeta     = os.path.join(
        CONFIG["carpeta_base"],
        f"{CLIENTE['nit']}_{CLIENTE['nombre']}",
        str(ultimo_dia.year),
        periodo
    )
    os.makedirs(carpeta, exist_ok=True)
    carpeta_xml = os.path.join(carpeta, "xml")

    xmls = sum(len([f for f in files if f.lower().endswith(".xml")])
               for _, _, files in os.walk(carpeta_xml))

    if xmls == 0:
        print("\n  Conectando al SAT...")
        MAX_INTENTOS = 3
        for intento in range(1, MAX_INTENTOS+1):
            try:
                if intento > 1: print(f"  🔄 Reintento {intento}/{MAX_INTENTOS}...")
                with sync_playwright() as p:
                    nav   = p.chromium.launch(channel="msedge", headless=False)
                    pag   = nav.new_page()
                    login(pag, CLIENTE["nit"], CLIENTE["password"])
                    print("  ✅ Login exitoso")
                    abrir_consulta_dte(pag)
                    frame = obtener_iframe(pag)
                    if not frame:
                        print("  ERROR: No se encontro iframe"); nav.close(); return
                    frame.wait_for_load_state("networkidle"); pag.wait_for_timeout(2000)
                    xmls = descargar_xmls(pag, frame, primer_dia, ultimo_dia, carpeta)
                    nav.close()
                break
            except Exception as e:
                print(f"  ⚠️  Intento {intento}: {str(e)[:80]}")
                if intento == MAX_INTENTOS:
                    print("  ❌ No se pudo conectar al SAT"); return
                import time; print("  ⏳ Esperando 15 segundos..."); time.sleep(15)
    else:
        print(f"\n  ℹ️  Usando {xmls} XMLs existentes")

    if xmls == 0:
        print("  ⚠️  Sin documentos para el día anterior"); return

    print("\n  Procesando facturas...")
    filas = procesar_xmls(carpeta_xml, periodo)
    print(f"  Facturas procesadas: {len(filas)}")
    if not filas: print("  ⚠️  Sin facturas procesadas"); return

    tipos = Counter(f["tipo_trans"] for f in filas)
    total = round(sum(f["gran_total"] for f in filas),2)
    print(f"  Tipos: M={tipos.get('M',0)} S={tipos.get('S',0)} C={tipos.get('C',0)}")
    print(f"  Total: Q {total:,.2f}")

    # Generar archivos
    print("\n  Generando archivos...")
    nb         = f"viaticos_{CLIENTE['nit']}_{periodo}"
    ruta_csv   = os.path.join(carpeta, f"{nb}.csv")
    ruta_excel = os.path.join(carpeta, f"{nb}.xlsx")
    titulo     = f"VIATICOS — {CLIENTE['nombre']} — {periodo}"
    generar_csv(filas, ruta_csv)
    generar_excel(filas, ruta_excel, titulo)

    # Subir a Google Sheets
    print("\n  Subiendo a Google Sheets...")
    subir_documentos_sheets(filas)

    # Correo
    print("\n  Enviando correo...")
    enviar_correo(ruta_excel, ruta_csv, filas, periodo)

    print("\n" + "="*55)
    print("  PROCESO COMPLETADO")
    print(f"  Archivos en: {carpeta}")
    print("="*55)

main()
