# ============================================================
#  APP VIÁTICOS — Formulario de Liquidación v2
#  Streamlit + Google Sheets
# ============================================================

import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
import xml.etree.ElementTree as ET
import os
import glob
from datetime import datetime
import unicodedata
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# ─────────────────────────────────────────────────────
# CONFIGURACION
# ─────────────────────────────────────────────────────
CREDENTIALS_FILE = "google_credentials.json"
SHEET_ID         = "13gJPXnQMaZNj5qFiUBUrZR8yiqG-w_8hWDmp-dW0jZI"
CARPETA_XML      = os.path.join(
    os.path.expanduser("~"),
    "Desktop", "agente-sat", "app_viaticos"
)

CATEGORIAS_GASTO = [
    "-- Selecciona una categoría --",
    "00-Combustible",
    "01-Hospedaje",
    "03-Datos",
    "04-Copias",
    "05-Pasajes",
    "06-Otros",
    "07-Viaticos",
    "08-Parqueos",
    "09-Varios",
]

SEMANAS = [f"Semana {i:02d}" for i in range(1, 53)]

# Mapeo de categorias a columnas del reporte
COLUMNAS_REPORTE = {
    "00-Combustible" : "Gasolina",
    "01-Hospedaje"   : "Hospedaje",
    "02-Alimentacion": "Alimentacion",
    "03-Datos"       : "Datos",
    "04-Copias"      : "Otros",
    "05-Pasajes"     : "Otros",
    "06-Otros"       : "Otros",
    "07-Viaticos"    : "Alimentacion",
    "08-Parqueos"    : "Otros",
    "09-Varios"      : "Otros",
}
# Orden de columnas segun formato fisico
COLS_REPORTE = ["Hospedaje","Alimentacion","Gasolina","Datos","Otros"]

NS_LIST = [
    {"dte": "http://www.sat.gob.gt/dte/fel/0.2.0"},
    {"dte": "http://www.sat.gob.gt/dte/fel/0.1.0"},
]

# ─────────────────────────────────────────────────────
# ESTILOS — visibles en modo claro y oscuro
# ─────────────────────────────────────────────────────
CSS = """
<style>
    .card {
        background: #1a2a3a;
        color: #ffffff;
        border-left: 4px solid #2E75B6;
        border-radius: 10px;
        padding: 1.2rem 1.4rem;
        margin: 0.8rem 0;
    }
    .card b { color: #90CAF9; }
    .card .monto {
        font-size: 1.3rem;
        font-weight: 700;
        color: #4FC3F7;
        margin-top: 0.5rem;
    }
    .card-comb {
        background: #1a3a1a;
        color: #ffffff;
        border-left: 4px solid #43A047;
        border-radius: 10px;
        padding: 1rem 1.4rem;
        margin: 0.8rem 0;
    }
    .card-comb b { color: #A5D6A7; }
    .card-ok {
        background: #1a3a1a;
        color: #ffffff;
        border-left: 4px solid #43A047;
        border-radius: 10px;
        padding: 1.2rem 1.4rem;
        margin: 0.8rem 0;
    }
    .card-ok h3 { color: #69F0AE; margin:0 0 0.5rem 0; }
    .card-ok b  { color: #A5D6A7; }
    .card-gsf {
        background: #2a1a3a;
        color: #ffffff;
        border-left: 4px solid #AB47BC;
        border-radius: 10px;
        padding: 1rem 1.4rem;
        margin: 0.8rem 0;
    }
    .card-gsf b { color: #CE93D8; }
    .badge {
        display: inline-block;
        padding: 2px 10px;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
        margin: 2px;
    }
    .badge-c { background:#1B5E20; color:#A5D6A7; }
    .badge-m { background:#3E2723; color:#BCAAA4; }
    .badge-s { background:#1A237E; color:#90CAF9; }
    .badge-s-ok { background:#1B5E20; color:#A5D6A7; }
    .badge-n { background:#B71C1C; color:#EF9A9A; }
    .badge-p { background:#F57F17; color:#FFF176; }
    .stButton>button {
        width: 100%;
        font-weight: bold;
        border-radius: 8px;
        padding: 0.6rem;
    }
</style>
"""

# ─────────────────────────────────────────────────────
# GOOGLE SHEETS
# ─────────────────────────────────────────────────────
@st.cache_resource
def conectar_sheets():
    # ── Modo nube: st.secrets ─────────────────────────
    try:
        secrets = st.secrets.get("gcp_service_account", None)
        if secrets is not None:
            creds_dict = {
                "type"                        : secrets.get("type","service_account"),
                "project_id"                  : secrets.get("project_id",""),
                "private_key_id"              : secrets.get("private_key_id",""),
                "private_key"                 : secrets.get("private_key","").replace("\\n","\n"),
                "client_email"                : secrets.get("client_email",""),
                "client_id"                   : secrets.get("client_id",""),
                "auth_uri"                    : secrets.get("auth_uri","https://accounts.google.com/o/oauth2/auth"),
                "token_uri"                   : secrets.get("token_uri","https://oauth2.googleapis.com/token"),
                "auth_provider_x509_cert_url" : "https://www.googleapis.com/oauth2/v1/certs",
                "client_x509_cert_url"        : f"https://www.googleapis.com/robot/v1/metadata/x509/{secrets.get('client_email','').replace('@','%40')}",
                "universe_domain"             : "googleapis.com"
            }
            creds  = Credentials.from_service_account_info(
                creds_dict,
                scopes=[
                    "https://spreadsheets.google.com/feeds",
                    "https://www.googleapis.com/auth/drive",
                    "https://www.googleapis.com/auth/spreadsheets",
                ]
            )
            return gspread.authorize(creds)
    except Exception as e:
        st.warning(f"⚠️ Modo nube falló: {e}")

    # ── Modo local: archivo JSON ──────────────────────
    try:
        if os.path.exists(CREDENTIALS_FILE):
            return gspread.service_account(filename=CREDENTIALS_FILE)
        else:
            st.error("⚠️ No se encontraron credenciales. Configura los Secrets en Streamlit Cloud.")
            return None
    except Exception as e:
        st.error(f"Error conectando Google Sheets: {e}")
        return None

def obtener_hoja(client):
    try:
        libro = client.open_by_key(SHEET_ID)
        hoja  = libro.get_worksheet(0)
        if not hoja.cell(1, 1).value:
            hoja.append_row([
                "Timestamp","Semana","CorreoEncuestador","TieneFact",
                "Correlativo_Interno","Fecha","HoraEmision","Serie",
                "Correlativo","Referencia","NombreEmisor","NIT",
                "Departamento","Municipio","Direccion",
                "MontoTotal","MontoIVA","MontoExento","IVADeducible",
                "TipoTransaccion","ClasificacionGasto","DescripcionCorta",
                "Galonaje","PrecioGalon","Estado"
            ])
        return hoja
    except Exception as e:
        st.error(f"Error Google Sheet: {type(e).__name__}: {e}")
        return None

def ya_liquidado(hoja, serie, correlativo):
    """Verifica duplicados por Serie, Correlativo o Referencia"""
    try:
        referencia = f"{serie}-{correlativo}"
        for r in hoja.get_all_records():
            # Verificar por serie + correlativo
            match_serie = (str(r.get("Serie","")) == str(serie) and
                          str(r.get("Correlativo","")) == str(correlativo))
            # Verificar por referencia (campo combinado)
            match_ref   = str(r.get("Referencia","")) == referencia
            # Verificar por Correlativo_Interno que contiene la referencia
            match_int   = str(r.get("Correlativo_Interno","")) == referencia

            if match_serie or match_ref or match_int:
                return True, r.get("CorreoEncuestador","")
    except:
        pass
    return False, ""

def siguiente_gsf(hoja):
    try:
        registros = hoja.get_all_records()
        nums = []
        for r in registros:
            ci = str(r.get("Correlativo_Interno",""))
            if ci.startswith("GSF-"):
                try:
                    nums.append(int(ci.split("-")[1]))
                except:
                    pass
        siguiente = max(nums) + 1 if nums else 1
        return f"GSF-{siguiente:03d}"
    except:
        return "GSF-001"

# ─────────────────────────────────────────────────────
# PARSEAR XMLs
# ─────────────────────────────────────────────────────
def limpiar(valor):
    if not valor: return ""
    texto = unicodedata.normalize("NFD", str(valor))
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    for car in [",",";",'"',"'","|","\n","\r","\t"]:
        texto = texto.replace(car, " ")
    while "  " in texto: texto = texto.replace("  ", " ")
    return texto.strip()

def limpiar_desc(desc):
    if not desc: return ""
    desc = desc.strip()
    if "|" in desc:
        partes = [p.strip() for p in desc.split("|")]
        p0 = partes[0].strip()
        if p0.isdigit() or (len(p0)<20 and p0.replace("-","").replace("_","").isalnum()):
            desc = "|".join(partes[1:]).strip()
    return limpiar(desc[:60])

def find_ns(root, path):
    for ns in NS_LIST:
        el = root.find(path, ns)
        if el is not None: return el, ns
    return None, None

def get_nc(imp, ns):
    nc = imp.find("dte:NombreCorto", ns)
    if nc is not None and nc.text: return nc.text.strip().lower()
    return (imp.get("NombreCorto","") or "").lower()

def extraer_combustible(item, ns):
    desc_el  = item.find("dte:Descripcion", ns)
    cant_el  = item.find("dte:Cantidad", ns)
    prec_el  = item.find("dte:PrecioUnitario", ns)
    desc     = (desc_el.text or "") if desc_el is not None else ""
    galones  = 0.0
    precio   = 0.0
    tipo_c   = ""
    if cant_el  is not None and cant_el.text:
        try: galones = round(float(cant_el.text), 4)
        except: pass
    if prec_el  is not None and prec_el.text:
        try: precio = round(float(prec_el.text), 4)
        except: pass
    if "|" in desc:
        partes = [p.strip() for p in desc.split("|")]
        tipo_c = partes[0]
        try:
            if len(partes) > 2: precio = max(precio, float(partes[2]))
        except: pass
    else:
        tipo_c = desc[:20]
    t = tipo_c.lower()
    if "regular" in t: tipo_c = "Regular"
    elif "super" in t: tipo_c = "Super"
    elif "diesel" in t: tipo_c = "Diesel"
    elif "premium" in t: tipo_c = "Premium"
    return tipo_c, round(galones, 4), round(precio, 4)

@st.cache_data(ttl=300, show_spinner="Cargando documentos...")
def cargar_xmls():
    """
    Primero intenta leer documentos desde Google Sheets (hoja Documentos).
    Si no encuentra, intenta leer XMLs locales como fallback.
    """
    xmls = {}

    # Modo nube/hibrido: leer desde Google Sheets
    try:
        client = conectar_sheets()
        if client:
            libro = client.open_by_key(SHEET_ID)
            hojas = [h.title for h in libro.worksheets()]
            if "Documentos" in hojas:
                hoja_docs = libro.worksheet("Documentos")
                registros = hoja_docs.get_all_records()
                for r in registros:
                    serie = str(r.get("Serie",""))
                    if not serie:
                        continue
                    xmls[serie] = {
                        "serie"        : serie,
                        "correlativo"  : str(r.get("Correlativo","")),
                        "referencia"   : str(r.get("Referencia","")),
                        "fecha"        : str(r.get("Fecha","")),
                        "hora"         : str(r.get("Hora","")),
                        "nit"          : str(r.get("NIT","")),
                        "emisor"       : str(r.get("Emisor","")),
                        "direccion"    : str(r.get("Direccion","")),
                        "municipio"    : str(r.get("Municipio","")),
                        "departamento" : str(r.get("Departamento","")),
                        "gran_total"   : float(r.get("MontoTotal",0) or 0),
                        "monto_iva"    : float(r.get("MontoIVA",0) or 0),
                        "monto_exento" : float(r.get("MontoExento",0) or 0),
                        "iva_ded"      : str(r.get("IVADeducible","S")),
                        "tipo_trans"   : str(r.get("TipoTransaccion","S")),
                        "desc_corta"   : str(r.get("DescCorta","")),
                        "es_combustible": str(r.get("EsCombustible","")) == "Si",
                        "tipo_comb"    : str(r.get("TipoComb","")),
                        "galones"      : float(r.get("Galones",0) or 0),
                        "precio_galon" : float(r.get("PrecioGalon",0) or 0),
                    }
                if xmls:
                    return xmls
    except Exception:
        pass

    # Fallback: leer XMLs locales
    archivos = glob.glob(os.path.join(CARPETA_XML,"**","xml","**","*.xml"), recursive=True)
    if not archivos:
        archivos = glob.glob(os.path.join(CARPETA_XML,"**","*.xml"), recursive=True)

    for ruta in archivos:
        try:
            tree = ET.parse(ruta)
            root = tree.getroot()
            _, ns = find_ns(root, ".//dte:DatosGenerales")
            if ns is None: continue
            dg      = root.find(".//dte:DatosGenerales", ns)
            num_aut = root.find(".//dte:NumeroAutorizacion", ns)
            emisor  = root.find(".//dte:Emisor", ns)
            totales = root.find(".//dte:Totales", ns)
            if dg is None or num_aut is None: continue
            serie       = num_aut.get("Serie","")
            correlativo = num_aut.get("Numero","")
            tipo_dte    = dg.get("Tipo","FACT")
            fecha_hora  = dg.get("FechaHoraEmision","")
            fecha = hora = ""
            try:
                dt    = datetime.strptime(fecha_hora[:19], "%Y-%m-%dT%H:%M:%S")
                fecha = dt.strftime("%d/%m/%Y")
                hora  = dt.strftime("%H:%M:%S")
            except:
                fecha = fecha_hora[:10]
                hora  = fecha_hora[11:19] if len(fecha_hora) > 10 else ""
            nit_emisor = nom_emisor = dir_emisor = mun_emisor = dep_emisor = ""
            if emisor is not None:
                nit_emisor = emisor.get("NITEmisor","") or emisor.get("Nit","")
                nom_emisor = emisor.get("NombreEmisor","") or emisor.get("NombreComercial","")
                dir_el = emisor.find("dte:DireccionEmisor", ns)
                if dir_el is not None:
                    d = dir_el.find("dte:Direccion",    ns); m = dir_el.find("dte:Municipio",    ns)
                    e = dir_el.find("dte:Departamento", ns)
                    dir_emisor = d.text if d is not None else ""
                    mun_emisor = m.text if m is not None else ""
                    dep_emisor = e.text if e is not None else ""
            gran_total = monto_iva = monto_exento = 0.0
            if totales is not None:
                gt = totales.find("dte:GranTotal", ns)
                if gt is not None and gt.text: gran_total = round(float(gt.text), 2)
                for timp in totales.findall("dte:TotalImpuestos/dte:TotalImpuesto", ns):
                    nc_el = timp.find("dte:NombreCorto", ns)
                    nc    = nc_el.text if nc_el is not None else timp.get("NombreCorto","")
                    if (nc or "").upper() == "IVA":
                        try:
                            tm = timp.find("dte:TotalMontoImpuesto", ns)
                            v  = tm.text if tm is not None else timp.get("TotalMontoImpuesto",0)
                            monto_iva = round(float(v), 2)
                        except: pass
            items   = root.findall(".//dte:Items/dte:Item", ns)
            es_comb = False; tipo_comb_str = ""; galones = precio_galon = 0.0
            for item in items:
                desc_el = item.find("dte:Descripcion", ns)
                desc    = (desc_el.text or "") if desc_el is not None else ""
                imps    = item.findall("dte:Impuestos/dte:Impuesto", ns)
                for imp in imps:
                    if "idp" in get_nc(imp, ns): es_comb = True; break
                if es_comb:
                    tipo_comb_str, galones, precio_galon = extraer_combustible(item, ns); break
                if any(p in desc.lower() for p in ["regular","super","diesel","premium","gasolina","combustible","galones"]):
                    es_comb = True
                    tipo_comb_str, galones, precio_galon = extraer_combustible(item, ns); break
            tipo_trans = "C" if es_comb else ("M" if any(i.get("BienOServicio","B")=="B" for i in items) else "S")
            iva_ded    = "P" if tipo_dte in ["FPEQ","FESP"] else ("N" if monto_iva==0 else "S")
            desc_corta = ""
            if items:
                de = items[0].find("dte:Descripcion", ns)
                if de is not None and de.text: desc_corta = limpiar_desc(de.text)
            if es_comb: desc_corta = f"Combustible {tipo_comb_str}".strip()
            xmls[serie] = {
                "serie":serie,"correlativo":correlativo,"referencia":f"{serie}-{correlativo}",
                "fecha":fecha,"hora":hora,"nit":limpiar(nit_emisor),"emisor":limpiar(nom_emisor),
                "direccion":limpiar(dir_emisor),"municipio":limpiar(mun_emisor),
                "departamento":limpiar(dep_emisor),"gran_total":gran_total,
                "monto_iva":monto_iva,"monto_exento":monto_exento,"iva_ded":iva_ded,
                "tipo_trans":tipo_trans,"desc_corta":desc_corta,"es_combustible":es_comb,
                "tipo_comb":tipo_comb_str,"galones":galones,"precio_galon":precio_galon,
            }
        except: pass
    return xmls

# ─────────────────────────────────────────────────────
# INTERFAZ
# ─────────────────────────────────────────────────────
def generar_reporte_excel(correo, semana, registros):
    """
    Genera Excel con formato de liquidacion:
    - Una fila por dia
    - Columnas: Fecha | Hospedaje | Alimentacion | Gasolina | Datos | Otros | Total Dia
    - Totales por columna al final
    - Observaciones de gastos sin factura
    - Cuadro de liquidacion: Recibido - Gastos = Resultado
    """
    from collections import defaultdict

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidacion"

    # ── Estilos ──────────────────────────────────────
    azul_osc = PatternFill("solid", fgColor="1F4E79")
    azul_med = PatternFill("solid", fgColor="2E75B6")
    azul_cla = PatternFill("solid", fgColor="DEEAF1")
    verde    = PatternFill("solid", fgColor="E2EFDA")
    gris_alt = PatternFill("solid", fgColor="F5F5F5")
    rojo_fill= PatternFill("solid", fgColor="C00000")
    borde    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    fmt_q = '#,##0.00'

    NCOLS = len(COLS_REPORTE) + 2  # Fecha + categorias + Total

    def celda(ws, row, col, val="", bold=False, color="000000",
              fill=None, fmt=None, align="left", wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=bold, color=color, size=10)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border    = borde
        if fill: c.fill = fill
        if fmt:  c.number_format = fmt
        return c

    def merge_hdr(ws, row, c1, c2, txt, fill=azul_osc):
        ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
        c = ws.cell(row=row, column=c1, value=txt)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Titulo ────────────────────────────────────────
    merge_hdr(ws, 1, 1, NCOLS, "LIQUIDACIÓN DE GASTOS")
    ws.row_dimensions[1].height = 30

    # Info encabezado
    merge_hdr(ws, 2, 1, 3, f"Encuestador:", azul_cla)
    ws.cell(row=2, column=1).font = Font(bold=True, size=10)
    merge_hdr(ws, 2, 4, NCOLS-2, correo, azul_cla)
    ws.cell(row=2, column=4).font = Font(size=10, color="000000")
    ws.cell(row=2, column=4).alignment = Alignment(horizontal="left")

    merge_hdr(ws, 3, 1, 2, "Semana:", azul_cla)
    ws.cell(row=3, column=1).font = Font(bold=True, size=10)
    ws.cell(row=3, column=3, value=semana).font = Font(size=10)
    ws.cell(row=3, column=3).alignment = Alignment(horizontal="left")
    merge_hdr(ws, 3, NCOLS-1, NCOLS,
              f"Generado: {datetime.now().strftime('%d/%m/%Y')}", azul_cla)
    ws.cell(row=3, column=NCOLS-1).font = Font(size=10, color="000000")

    # ── Encabezados tabla ─────────────────────────────
    encabezados = ["Fecha"] + COLS_REPORTE + ["Total Día"]
    for col, h in enumerate(encabezados, 1):
        celda(ws, 4, col, h, bold=True, color="FFFFFF",
              fill=azul_med, align="center")
    ws.row_dimensions[4].height = 28

    # ── Agrupar por fecha ─────────────────────────────
    por_dia = defaultdict(lambda: {c: 0.0 for c in COLS_REPORTE})
    obs_list = []
    total_gral = 0.0

    for r in registros:
        cat   = r.get("ClasificacionGasto","")
        col_c = COLUMNAS_REPORTE.get(cat, "Otros")
        monto = float(r.get("MontoTotal", 0) or 0)
        fecha = r.get("Fecha","Sin fecha")
        tiene = r.get("TieneFact","Sí")
        desc  = r.get("DescripcionCorta","")
        emisor= r.get("NombreEmisor","") or desc

        por_dia[fecha][col_c] += monto
        total_gral += monto

        if tiene == "No" and desc:
            obs_list.append(f"{fecha} | {cat} | {emisor} | Q{monto:,.2f}")

    # Ordenar dias cronologicamente
    def sort_fecha(f):
        try:
            return datetime.strptime(f, "%d/%m/%Y")
        except:
            return datetime.min

    dias_ordenados = sorted(por_dia.keys(), key=sort_fecha)

    # ── Filas por dia ─────────────────────────────────
    totales_col = {c: 0.0 for c in COLS_REPORTE}
    for ri, fecha in enumerate(dias_ordenados):
        fila   = 5 + ri
        fill_f = gris_alt if ri % 2 == 0 else None
        celda(ws, fila, 1, fecha, fill=fill_f)
        total_dia = 0.0
        for ci, col_n in enumerate(COLS_REPORTE, 2):
            val = por_dia[fecha][col_n]
            c   = celda(ws, fila, ci,
                        val if val > 0 else None,
                        fill=fill_f, fmt=fmt_q, align="right")
            if val > 0: c.fill = verde
            totales_col[col_n] += val
            total_dia += val
        celda(ws, fila, NCOLS, total_dia, bold=True,
              fill=azul_cla, fmt=fmt_q, align="right")

    # ── Fila totales ──────────────────────────────────
    fila_tot = 5 + len(dias_ordenados)
    ws.merge_cells(f"A{fila_tot}:A{fila_tot}")
    celda(ws, fila_tot, 1, "TOTALES", bold=True,
          color="FFFFFF", fill=azul_med, align="center")
    for ci, col_n in enumerate(COLS_REPORTE, 2):
        celda(ws, fila_tot, ci, totales_col[col_n],
              bold=True, color="FFFFFF", fill=azul_med,
              fmt=fmt_q, align="right")
    celda(ws, fila_tot, NCOLS, total_gral,
          bold=True, color="FFFFFF", fill=azul_osc,
          fmt=fmt_q, align="right")
    ws.row_dimensions[fila_tot].height = 22

    # ── Observaciones ─────────────────────────────────
    fila_obs = fila_tot + 2
    merge_hdr(ws, fila_obs, 1, NCOLS,
              "OBSERVACIONES / COMENTARIOS", azul_med)
    ws.row_dimensions[fila_obs].height = 22
    if obs_list:
        for i, obs in enumerate(obs_list):
            fo = fila_obs + 1 + i
            ws.merge_cells(f"A{fo}:{get_column_letter(NCOLS)}{fo}")
            ws.cell(row=fo, column=1, value=f"• {obs}")
            ws.row_dimensions[fo].height = 16
        fila_cuadro = fila_obs + len(obs_list) + 2
    else:
        # Espacio en blanco para escribir
        for i in range(3):
            fo = fila_obs + 1 + i
            ws.merge_cells(f"A{fo}:{get_column_letter(NCOLS)}{fo}")
            ws.cell(row=fo, column=1).border = borde
            ws.row_dimensions[fo].height = 18
        fila_cuadro = fila_obs + 5

    # ── Cuadro de liquidacion ─────────────────────────
    # Viáticos recibidos
    merge_hdr(ws, fila_cuadro, 1, NCOLS, "RESUMEN DE LIQUIDACIÓN", azul_osc)
    ws.row_dimensions[fila_cuadro].height = 22

    fila_rec = fila_cuadro + 1
    ws.merge_cells(f"A{fila_rec}:{get_column_letter(NCOLS-2)}{fila_rec}")
    celda(ws, fila_rec, 1, "Viáticos recibidos (Q):",
          bold=True, fill=azul_cla, align="left")
    celda(ws, fila_rec, NCOLS-1, "", fill=azul_cla)  # espacio editable
    celda(ws, fila_rec, NCOLS, "", fmt=fmt_q, align="right")

    fila_gas = fila_rec + 1
    ws.merge_cells(f"A{fila_gas}:{get_column_letter(NCOLS-2)}{fila_gas}")
    celda(ws, fila_gas, 1, "Total gastos liquidados (Q):",
          bold=True, fill=azul_cla, align="left")
    celda(ws, fila_gas, NCOLS, total_gral,
          bold=True, fill=azul_cla, fmt=fmt_q, align="right")

    fila_sal = fila_gas + 1
    ws.merge_cells(f"A{fila_sal}:{get_column_letter(NCOLS-2)}{fila_sal}")
    celda(ws, fila_sal, 1, "Resultado  (Recibido - Gastos):",
          bold=True, color="FFFFFF", fill=azul_osc, align="left")
    celda(ws, fila_sal, NCOLS, "",
          bold=True, color="FFFFFF", fill=azul_osc,
          fmt=fmt_q, align="right")

    # Nota
    fila_nota = fila_sal + 1
    ws.merge_cells(f"A{fila_nota}:{get_column_letter(NCOLS)}{fila_nota}")
    c = ws.cell(row=fila_nota, column=1,
                value="* Complete el campo 'Viáticos recibidos' para calcular el resultado")
    c.font      = Font(italic=True, size=9, color="666666")
    c.alignment = Alignment(horizontal="left")

    # ── Firmas ────────────────────────────────────────
    fila_f = fila_nota + 2
    ws.merge_cells(f"A{fila_f}:{get_column_letter(NCOLS//3)}{fila_f}")
    ws.cell(row=fila_f, column=1,
            value="Firma encuestador: _______________")
    ws.merge_cells(f"{get_column_letter(NCOLS//3+2)}{fila_f}:{get_column_letter(NCOLS//3*2)}{fila_f}")
    ws.cell(row=fila_f, column=NCOLS//3+2,
            value="Revisión 1: _______________")
    ws.merge_cells(f"{get_column_letter(NCOLS//3*2+2)}{fila_f}:{get_column_letter(NCOLS)}{fila_f}")
    ws.cell(row=fila_f, column=NCOLS//3*2+2,
            value="Revisión 2: _______________")
    ws.row_dimensions[fila_f].height = 28

    # ── Detalle de facturas liquidadas ────────────────
    fila_det = fila_f + 2
    merge_hdr(ws, fila_det, 1, NCOLS,
              "DETALLE DE DOCUMENTOS LIQUIDADOS", azul_osc)
    ws.row_dimensions[fila_det].height = 22

    # Encabezados detalle
    cols_det = ["#","Fecha","Referencia","Nombre Emisor","Tipo de Gasto","Monto (Q)"]
    anchos_det = [4, 12, 24, 38, 20, 12]
    fila_det_hdr = fila_det + 1
    for ci, (h, aw) in enumerate(zip(cols_det, anchos_det), 1):
        celda(ws, fila_det_hdr, ci, h,
              bold=True, color="FFFFFF", fill=azul_med,
              align="center")
    ws.row_dimensions[fila_det_hdr].height = 22

    # Ordenar registros por fecha
    regs_ord = sorted(registros, key=lambda r: sort_fecha(r.get("Fecha","")))

    total_det = 0.0
    for di, r in enumerate(regs_ord, 1):
        fila_di  = fila_det_hdr + di
        fecha_d  = r.get("Fecha","")
        nit_d    = r.get("NIT","") or r.get("Nit","")
        emisor_d = r.get("NombreEmisor","") or r.get("DescripcionCorta","Sin factura")
        tipo_d   = r.get("ClasificacionGasto","")
        monto_d  = float(r.get("MontoTotal",0) or 0)
        fact_d   = r.get("TieneFact","Sí")
        corr_d   = r.get("Correlativo_Interno","") or r.get("Referencia","")

        fill_di = gris_alt if di % 2 == 0 else None

        # Referencia: usar la del doc o GSF si sin factura
        ref_d = corr_d if fact_d == "No" else r.get("Referencia","")

        celda(ws, fila_di, 1, di,       fill=fill_di, align="center")
        celda(ws, fila_di, 2, fecha_d,  fill=fill_di, align="center")
        c_ref = celda(ws, fila_di, 3, ref_d, fill=fill_di, align="left")
        if fact_d == "No":
            c_ref.fill = PatternFill("solid", fgColor="FFF2CC")
        celda(ws, fila_di, 4, emisor_d, fill=fill_di, align="left")
        celda(ws, fila_di, 5, tipo_d,   fill=fill_di, align="left")
        celda(ws, fila_di, 6, monto_d,
              fill=fill_di, fmt=fmt_q, align="right")

        total_det += monto_d

    # Total detalle
    fila_tot_det = fila_det_hdr + len(regs_ord) + 1
    ws.merge_cells(f"A{fila_tot_det}:E{fila_tot_det}")
    celda(ws, fila_tot_det, 1, f"TOTAL — {len(regs_ord)} documentos",
          bold=True, color="FFFFFF", fill=azul_osc, align="right")
    celda(ws, fila_tot_det, 6, total_det,
          bold=True, color="FFFFFF", fill=azul_osc,
          fmt=fmt_q, align="right")

    # ── Anchos de columnas ────────────────────────────
    # Columnas del reporte principal
    anchos = [12] + [14] * len(COLS_REPORTE) + [14]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # Ajustar columna emisor (col 4) para detalle
    ws.column_dimensions["D"].width = 40
    ws.freeze_panes = "B5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, total_gral, totales_col


def generar_reporte_pdf(correo, semana, registros, total_gral, totales_col):
    """Genera PDF del reporte de liquidacion"""
    from collections import defaultdict

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(10, 10, 10)

    # Colores
    AZUL_OSC = (31,  78, 121)
    AZUL_MED = (46, 117, 182)
    AZUL_CLA = (222, 234, 246)
    VERDE    = (226, 239, 218)
    GRIS     = (245, 245, 245)
    BLANCO   = (255, 255, 255)

    def rect_fill(x, y, w, h, color):
        pdf.set_fill_color(*color)
        pdf.rect(x, y, w, h, "F")

    def txt(x, y, w, h, texto, size=9, bold=False,
            color=(0,0,0), bg=None, align="L", border=0):
        if bg: pdf.set_fill_color(*bg)
        pdf.set_text_color(*color)
        pdf.set_font("Helvetica", "B" if bold else "", size)
        pdf.set_xy(x, y)
        pdf.cell(w, h, str(texto)[:60], border=border,
                 align=align, fill=bool(bg))

    W = 277  # ancho util landscape A4
    x0 = 10

    # Titulo
    rect_fill(x0, 10, W, 10, AZUL_OSC)
    txt(x0, 10, W, 10, "LIQUIDACION DE GASTOS",
        size=13, bold=True, color=BLANCO, align="C")

    # Info encabezado
    rect_fill(x0, 22, W*0.5, 7, AZUL_CLA)
    rect_fill(x0+W*0.5, 22, W*0.25, 7, AZUL_CLA)
    rect_fill(x0+W*0.75, 22, W*0.25, 7, AZUL_CLA)
    txt(x0, 22, W*0.5, 7, f"Encuestador: {correo}", bold=True, color=(0,0,0))
    txt(x0+W*0.5, 22, W*0.25, 7, f"Semana: {semana}", bold=True, color=(0,0,0))
    txt(x0+W*0.75, 22, W*0.25, 7,
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        color=(0,0,0), align="R")

    # Tabla principal - encabezados
    y_tab = 32
    col_w = [20] + [32]*len(COLS_REPORTE) + [28]
    encabs = ["Fecha"] + [c[:10] for c in COLS_REPORTE] + ["Total Dia"]
    rect_fill(x0, y_tab, sum(col_w), 8, AZUL_MED)
    xi = x0
    for i, (h, w) in enumerate(zip(encabs, col_w)):
        txt(xi, y_tab, w, 8, h, bold=True, color=BLANCO, align="C", border=1)
        xi += w

    # Agrupar por dia
    por_dia = defaultdict(lambda: {c: 0.0 for c in COLS_REPORTE})
    for r in registros:
        cat   = r.get("ClasificacionGasto","")
        col_c = COLUMNAS_REPORTE.get(cat, "Otros")
        monto = float(r.get("MontoTotal",0) or 0)
        fecha = r.get("Fecha","Sin fecha")
        por_dia[fecha][col_c] += monto

    def sk(f):
        try: return datetime.strptime(f,"%d/%m/%Y")
        except: return datetime.min

    dias_ord = sorted(por_dia.keys(), key=sk)
    y_cur = y_tab + 8

    for ri, fecha in enumerate(dias_ord):
        bg = GRIS if ri%2==0 else BLANCO
        rect_fill(x0, y_cur, sum(col_w), 7, bg)
        xi = x0
        txt(xi, y_cur, col_w[0], 7, fecha, align="C", border=1); xi += col_w[0]
        total_d = 0.0
        for ci, cn in enumerate(COLS_REPORTE):
            val = por_dia[fecha][cn]
            total_d += val
            v_txt = f"{val:,.2f}" if val > 0 else ""
            bg2 = VERDE if val > 0 else bg
            rect_fill(xi, y_cur, col_w[ci+1], 7, bg2)
            txt(xi, y_cur, col_w[ci+1], 7, v_txt, align="R", border=1)
            xi += col_w[ci+1]
        txt(xi, y_cur, col_w[-1], 7, f"{total_d:,.2f}",
            bold=True, bg=AZUL_CLA, align="R", border=1)
        y_cur += 7

    # Fila totales
    rect_fill(x0, y_cur, sum(col_w), 8, AZUL_MED)
    txt(x0, y_cur, col_w[0], 8, "TOTALES",
        bold=True, color=BLANCO, align="C", border=1)
    xi = x0 + col_w[0]
    for ci, cn in enumerate(COLS_REPORTE):
        txt(xi, y_cur, col_w[ci+1], 8,
            f"{totales_col.get(cn,0):,.2f}",
            bold=True, color=BLANCO, align="R", border=1)
        xi += col_w[ci+1]
    txt(xi, y_cur, col_w[-1], 8, f"{total_gral:,.2f}",
        bold=True, color=BLANCO, bg=AZUL_OSC, align="R", border=1)
    y_cur += 12

    # Cuadro liquidacion
    rect_fill(x0, y_cur, W, 7, AZUL_OSC)
    txt(x0, y_cur, W, 7, "RESUMEN DE LIQUIDACION",
        bold=True, color=BLANCO, align="C")
    y_cur += 7
    w_lbl = W * 0.6
    w_val = W * 0.4

    rect_fill(x0, y_cur, w_lbl, 7, AZUL_CLA)
    txt(x0, y_cur, w_lbl, 7,
        "Viaticos autorizados/recibidos (Q):", bold=True, border=1)
    pdf.set_draw_color(46, 117, 182)
    pdf.rect(x0+w_lbl, y_cur, w_val, 7)
    y_cur += 7

    rect_fill(x0, y_cur, w_lbl, 7, AZUL_CLA)
    txt(x0, y_cur, w_lbl, 7, "Total gastos liquidados (Q):",
        bold=True, border=1)
    txt(x0+w_lbl, y_cur, w_val, 7, f"Q {total_gral:,.2f}",
        bold=True, bg=AZUL_CLA, align="R", border=1)
    y_cur += 7

    rect_fill(x0, y_cur, w_lbl, 7, AZUL_OSC)
    txt(x0, y_cur, w_lbl, 7, "Resultado (Recibido - Gastos):",
        bold=True, color=BLANCO, border=1)
    pdf.set_draw_color(31, 78, 121)
    pdf.rect(x0+w_lbl, y_cur, w_val, 7)
    y_cur += 12

    # Firmas
    fw = W / 3 - 5
    for i, lbl in enumerate(["Firma encuestador:", "Revision 1:", "Revision 2:"]):
        xi = x0 + i*(fw+5)
        txt(xi, y_cur, fw, 7, lbl, bold=True)
        pdf.line(xi, y_cur+12, xi+fw, y_cur+12)
    y_cur += 20

    # Detalle facturas - nueva pagina si no hay espacio
    if y_cur > 170:
        pdf.add_page()
        y_cur = 15

    rect_fill(x0, y_cur, W, 8, AZUL_OSC)
    txt(x0, y_cur, W, 8, "DETALLE DE DOCUMENTOS LIQUIDADOS",
        bold=True, color=BLANCO, align="C", size=10)
    y_cur += 8

    # Encabezados detalle
    dw = [8, 22, 55, 75, 45, 25, 47]
    dh_cols = ["#","Fecha","Referencia","Nombre Emisor","Tipo Gasto","Monto","Factura"]
    rect_fill(x0, y_cur, sum(dw), 7, AZUL_MED)
    xi = x0
    for h, w in zip(dh_cols, dw):
        txt(xi, y_cur, w, 7, h, bold=True, color=BLANCO, align="C", border=1)
        xi += w
    y_cur += 7

    regs_ord = sorted(registros, key=lambda r: sk(r.get("Fecha","")))
    total_det = 0.0
    for di, r in enumerate(regs_ord, 1):
        if y_cur > 190:
            pdf.add_page()
            y_cur = 15
        fecha_d  = r.get("Fecha","")
        fact_d   = r.get("TieneFact","Sí")
        corr_d   = r.get("Correlativo_Interno","") or r.get("Referencia","")
        ref_d    = corr_d if fact_d=="No" else r.get("Referencia","")
        emisor_d = (r.get("NombreEmisor","") or r.get("DescripcionCorta",""))[:40]
        tipo_d   = r.get("ClasificacionGasto","")
        monto_d  = float(r.get("MontoTotal",0) or 0)
        bg_d     = GRIS if di%2==0 else BLANCO
        if fact_d == "No":
            bg_d = (255, 242, 204)

        rect_fill(x0, y_cur, sum(dw), 6, bg_d)
        xi = x0
        vals = [str(di), fecha_d, ref_d, emisor_d, tipo_d,
                f"{monto_d:,.2f}", fact_d]
        aligns = ["C","C","L","L","L","R","C"]
        for v, w, al in zip(vals, dw, aligns):
            txt(xi, y_cur, w, 6, v, size=8, align=al, border=1)
            xi += w
        total_det += monto_d
        y_cur += 6

    # Total detalle
    rect_fill(x0, y_cur, sum(dw), 7, AZUL_OSC)
    txt(x0, y_cur, sum(dw)-dw[-1]-dw[-2], 7,
        f"TOTAL  {len(regs_ord)} documentos",
        bold=True, color=BLANCO, align="R", border=1)
    txt(x0+sum(dw)-dw[-1]-dw[-2], y_cur, dw[-2], 7,
        f"{total_det:,.2f}", bold=True, color=BLANCO,
        align="R", border=1)
    txt(x0+sum(dw)-dw[-1], y_cur, dw[-1], 7,
        "", bold=True, color=BLANCO, border=1)

    buf = io.BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    return buf


def main():
    st.set_page_config(
        page_title="App Viáticos — Aprende",
        page_icon="📋",
        layout="centered"
    )
    st.markdown(CSS, unsafe_allow_html=True)

    st.image("https://img.icons8.com/color/96/invoice.png", width=55)
    st.title("Liquidación de Viáticos")
    st.markdown("**Corporación Educativa en Informática — Aprende**")
    st.divider()

    # Pestañas
    tab1, tab2 = st.tabs(["📝 Registrar gasto", "📊 Mi reporte"])

    with tab2:
        st.subheader("📊 Reporte de liquidación semanal")
        st.markdown("Ingresa tu correo y semana para ver y descargar tu reporte.")

        col1, col2 = st.columns(2)
        with col1:
            correo_rep = st.text_input("Tu correo *", key="rep_correo",
                                       placeholder="ejemplo@grupoaprende.com")
        with col2:
            semana_rep = st.selectbox("Semana *", options=SEMANAS, key="rep_semana")

        if st.button("🔍 Ver mi reporte", type="primary", key="btn_reporte"):
            if not correo_rep or "@" not in correo_rep:
                st.error("⚠️ Ingresa un correo válido")
            else:
                with st.spinner("Buscando tus gastos..."):
                    client = conectar_sheets()
                    if client:
                        hoja      = obtener_hoja(client)
                        registros = hoja.get_all_records() if hoja else []
                        mis_gastos = [
                            r for r in registros
                            if str(r.get("CorreoEncuestador","")).lower() == correo_rep.lower()
                            and str(r.get("Semana","")) == semana_rep
                        ]

                if not mis_gastos:
                    st.warning(f"No se encontraron gastos para **{correo_rep}** en **{semana_rep}**")
                else:
                    st.success(f"✅ {len(mis_gastos)} gastos encontrados")

                    # Resumen por categoria
                    totales = {}
                    total_gral = 0.0
                    for r in mis_gastos:
                        cat   = r.get("ClasificacionGasto","Sin categoría")
                        monto = float(r.get("MontoTotal",0) or 0)
                        totales[cat] = totales.get(cat, 0) + monto
                        total_gral  += monto

                    # Mostrar resumen
                    cols = st.columns(min(len(totales), 4))
                    for i, (cat, monto) in enumerate(totales.items()):
                        with cols[i % len(cols)]:
                            st.metric(cat, f"Q {monto:,.2f}")

                    st.markdown(f"""
                    <div class="card">
                        <b>Total gastos {semana_rep}:</b>
                        <div class="monto">Q {total_gral:,.2f}</div>
                    </div>
                    """, unsafe_allow_html=True)

                    # Tabla detalle
                    st.subheader("Detalle de gastos")
                    import pandas as pd
                    df = pd.DataFrame([{
                        "Fecha"        : r.get("Fecha",""),
                        "Emisor"       : r.get("NombreEmisor","") or r.get("DescripcionCorta",""),
                        "Categoria"    : r.get("ClasificacionGasto",""),
                        "Tipo"         : r.get("TipoTransaccion",""),
                        "Factura"      : r.get("TieneFact",""),
                        "Monto"        : float(r.get("MontoTotal",0) or 0),
                    } for r in mis_gastos])
                    st.dataframe(df, use_container_width=True, hide_index=True)

                    # Descargar Excel
                    buf, total, tots = generar_reporte_excel(
                        correo_rep, semana_rep, mis_gastos
                    )
                    nombre_archivo = (f"Liquidacion_{correo_rep.split('@')[0]}"
                                     f"_{semana_rep.replace(' ','')}.xlsx")
                    col_xl, col_pdf = st.columns(2)
                    with col_xl:
                        st.download_button(
                            label="⬇️ Descargar Excel",
                            data=buf,
                            file_name=nombre_archivo,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )
                    with col_pdf:
                        try:
                            buf_pdf = generar_reporte_pdf(
                                correo_rep, semana_rep,
                                mis_gastos, total, tots
                            )
                            nombre_pdf = nombre_archivo.replace(".xlsx",".pdf")
                            st.download_button(
                                label="⬇️ Descargar PDF",
                                data=buf_pdf,
                                file_name=nombre_pdf,
                                mime="application/pdf",
                                type="secondary"
                            )
                        except Exception as ep:
                            st.warning(f"PDF no disponible: {ep}")

    with tab1:
        xmls = cargar_xmls()
        st.success(f"✅ {len(xmls)} documentos disponibles")
        st.divider()

    # ── DATOS DEL ENCUESTADOR ──
    st.subheader("👤 Datos del encuestador")
    col1, col2 = st.columns([3,2])
    with col1:
        correo = st.text_input(
            "Correo electrónico *",
            placeholder="ejemplo@grupoaprende.com"
        )
    with col2:
        semana = st.selectbox("Semana de liquidación *", options=SEMANAS)

    st.divider()

    # ── ¿TIENE FACTURA? ──
    st.subheader("🧾 Documento")
    tiene_fact = st.radio(
        "¿El gasto tiene factura?",
        options=["Sí, tiene factura", "No tiene factura"],
        horizontal=True
    )
    con_factura = tiene_fact == "Sí, tiene factura"

    doc = None
    gsf_code = None

    if con_factura:
        serie = st.text_input(
            "Serie del documento *",
            placeholder="Ej: F191A3D8"
        ).strip().upper()

        if st.button("🔍 Buscar documento", type="primary"):
            if not correo or "@" not in correo:
                st.error("⚠️ Ingresa un correo válido")
                st.stop()
            if not serie:
                st.error("⚠️ Ingresa la serie")
                st.stop()
            if serie not in xmls:
                st.error(f"❌ Serie **{serie}** no encontrada")
                st.info("💡 Verifica la serie en la parte superior de tu factura")
                st.stop()

            # Verificar si ya fue liquidado
            client = conectar_sheets()
            hoja   = obtener_hoja(client) if client else None
            if hoja:
                ya, quien = ya_liquidado(hoja, serie, xmls[serie]["correlativo"])
                if ya:
                    st.warning(f"⚠️ Este documento ya fue liquidado por **{quien}**")
                    st.stop()

            st.session_state["doc"]          = xmls[serie]
            st.session_state["correo"]       = correo
            st.session_state["semana"]       = semana
            st.session_state["con_factura"]  = True
            st.session_state["busqueda_ok"]  = True

    else:
        # Sin factura
        if st.button("➕ Registrar gasto sin factura", type="primary"):
            if not correo or "@" not in correo:
                st.error("⚠️ Ingresa un correo válido")
                st.stop()
            client = conectar_sheets()
            hoja   = obtener_hoja(client) if client else None
            gsf    = siguiente_gsf(hoja) if hoja else "GSF-001"
            st.session_state["gsf_code"]     = gsf
            st.session_state["correo"]       = correo
            st.session_state["semana"]       = semana
            st.session_state["con_factura"]  = False
            st.session_state["busqueda_ok"]  = True

    # ── MOSTRAR DETALLES ──
    if st.session_state.get("busqueda_ok"):
        correo    = st.session_state["correo"]
        semana    = st.session_state["semana"]
        con_fact  = st.session_state.get("con_factura", True)

        st.divider()

        if con_fact and "doc" in st.session_state:
            doc = st.session_state["doc"]

            tipo_badge = {
                "C": '<span class="badge badge-c">🔵 Combustible</span>',
                "M": '<span class="badge badge-m">🟤 Materiales</span>',
                "S": '<span class="badge badge-s">🟢 Servicio</span>',
            }.get(doc["tipo_trans"], doc["tipo_trans"])

            iva_badge = {
                "S": '<span class="badge badge-s-ok">✅ IVA Deducible</span>',
                "N": '<span class="badge badge-n">❌ No Deducible</span>',
                "P": '<span class="badge badge-p">🟡 Pequeño Contribuyente</span>',
            }.get(doc["iva_ded"], "")

            st.subheader("📄 Documento encontrado")
            st.markdown(f"""
            <div class="card">
                <b>Emisor:</b> {doc['emisor']}<br>
                <b>NIT:</b> {doc['nit']}<br>
                <b>Dirección:</b> {doc['direccion']}<br>
                <b>Municipio:</b> {doc['municipio']} &nbsp;
                <b>Departamento:</b> {doc['departamento']}<br>
                <b>Fecha:</b> {doc['fecha']} &nbsp;
                <b>Hora:</b> {doc['hora']}<br>
                <b>Serie-Correlativo:</b> {doc['referencia']}<br>
                <b>Descripción:</b> {doc['desc_corta']}<br>
                <br>{tipo_badge} &nbsp; {iva_badge}
                <div class="monto">Q {doc['gran_total']:,.2f}</div>
            </div>
            """, unsafe_allow_html=True)

            # Datos extra de combustible
            if doc["es_combustible"] and doc["galones"] > 0:
                st.markdown(f"""
                <div class="card-comb">
                    <b>⛽ Detalle de combustible</b><br>
                    <b>Tipo:</b> {doc['tipo_comb']}<br>
                    <b>Galones:</b> {doc['galones']:,.4f} Gl<br>
                    <b>Precio por galón:</b> Q {doc['precio_galon']:,.4f}
                </div>
                """, unsafe_allow_html=True)

        else:
            # Sin factura
            gsf_code = st.session_state.get("gsf_code","GSF-001")
            st.subheader("📋 Gasto sin factura")
            st.markdown(f"""
            <div class="card-gsf">
                <b>Correlativo interno generado:</b> {gsf_code}<br>
                <b>Encuestador:</b> {correo}<br>
                <b>Semana:</b> {semana}
            </div>
            """, unsafe_allow_html=True)

            col_m, col_o = st.columns([1, 2])
            with col_m:
                monto_gsf = st.number_input(
                    "Monto del gasto (Q) *",
                    min_value=0.0,
                    step=0.01,
                    format="%.2f",
                    help="Ingresa el monto total del gasto"
                )
            with col_o:
                observaciones_gsf = st.text_area(
                    "Observaciones *",
                    placeholder="Describe el gasto realizado...",
                    max_chars=200,
                    help="Detalla el concepto del gasto sin factura"
                )
            st.session_state["monto_gsf"]        = monto_gsf
            st.session_state["observaciones_gsf"] = observaciones_gsf

        # ── CLASIFICACION ──
        st.divider()
        st.subheader("🏷️ Clasificar gasto")

        clasificacion = st.selectbox(
            "Categoría de gasto *",
            options=CATEGORIAS_GASTO,
            index=0,
            help="Selecciona la categoría correcta"
        )

        if con_fact and doc and doc["es_combustible"]:
            st.info("💡 Esta factura contiene combustible. Considera seleccionar **00-Combustible**")

        # ── CONFIRMAR ──
        st.divider()
        if st.button("✅ Confirmar liquidación", type="primary"):
            # Validaciones
            if not correo or "@" not in correo:
                st.error("⚠️ Correo inválido"); st.stop()
            if clasificacion == "-- Selecciona una categoría --":
                st.error("⚠️ Debes seleccionar una categoría de gasto"); st.stop()
            if not con_fact:
                if st.session_state.get("monto_gsf", 0) <= 0:
                    st.error("⚠️ Ingresa el monto del gasto"); st.stop()
                if not st.session_state.get("observaciones_gsf","").strip():
                    st.error("⚠️ Ingresa las observaciones del gasto"); st.stop()

            with st.spinner("Registrando..."):
                client = conectar_sheets()
                if not client: st.stop()
                hoja = obtener_hoja(client)
                if not hoja: st.stop()

                # Doble verificacion al confirmar
                if con_fact and "doc" in st.session_state:
                    doc   = st.session_state["doc"]
                    ya2, quien2 = ya_liquidado(hoja, doc["serie"], doc["correlativo"])
                    if ya2:
                        st.warning(f"⚠️ Este documento ya fue liquidado por **{quien2}**. No se guardó.")
                        st.session_state.clear()
                        st.stop()
                    fila = [
                        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                        semana, correo, "Sí",
                        doc["referencia"],
                        doc["fecha"], doc["hora"],
                        doc["serie"], doc["correlativo"], doc["referencia"],
                        doc["emisor"], doc["nit"],
                        doc["departamento"], doc["municipio"], doc["direccion"],
                        doc["gran_total"], doc["monto_iva"], doc["monto_exento"],
                        doc["iva_ded"], doc["tipo_trans"],
                        clasificacion, doc["desc_corta"],
                        doc["galones"], doc["precio_galon"],
                        "Liquidado"
                    ]
                    emit  = doc["emisor"]
                    monto = doc["gran_total"]
                    ref   = doc["referencia"]
                else:
                    gsf_code      = st.session_state.get("gsf_code","GSF-001")
                    monto_gsf     = st.session_state.get("monto_gsf", 0.0)
                    obs_gsf       = st.session_state.get("observaciones_gsf","")
                    fila = [
                        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                        semana, correo, "No",
                        gsf_code,
                        datetime.now().strftime("%d/%m/%Y"),
                        datetime.now().strftime("%H:%M:%S"),
                        "","","",
                        "Sin factura","",
                        "","","",
                        monto_gsf, 0, 0, "", "",
                        clasificacion, obs_gsf,
                        0, 0,
                        "Liquidado"
                    ]
                    emit  = "Sin factura"
                    monto = monto_gsf
                    ref   = gsf_code

                hoja.append_row(fila)

            st.session_state.clear()
            st.markdown(f"""
            <div class="card-ok">
                <h3>✅ ¡Liquidación registrada!</h3>
                <b>Encuestador:</b> {correo}<br>
                <b>Semana:</b> {semana}<br>
                <b>Referencia:</b> {ref}<br>
                <b>Emisor:</b> {emit}<br>
                <b>Monto:</b> Q {monto:,.2f}<br>
                <b>Clasificación:</b> {clasificacion}<br>
                <b>Registro:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}
            </div>
            """, unsafe_allow_html=True)
            st.balloons()
            if st.button("➕ Liquidar otro gasto", type="primary"):
                st.rerun()

        st.divider()
        st.caption("App Viáticos v2 — ASPRO BPO | Corporación Educativa en Informática")

if __name__ == "__main__":
    main()
